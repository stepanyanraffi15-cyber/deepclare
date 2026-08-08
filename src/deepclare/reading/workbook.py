"""A7 — the workbook loader: one streaming pass, and the only read there will be.

A workbook's declared used range is routinely stale. Formatting applied down a whole
column inflates it to the sheet maximum, and a loader that materializes that range pays
for every phantom cell it names. So the sheet is opened in streaming mode *and* its
declared dimensions are discarded, which is what stops the reader padding every row out
to a width no data reaches.

A streaming sheet is forward-only. This pass is therefore the only read of the file, and
every later node — the whole-text render, the table locator, the column labeller, the
typed cell reader — works off the buffer it leaves behind rather than reopening anything.

What the buffer keeps, per non-empty row: each cell's typed number when it has one and
the text it renders as, indexed by column so that position survives. Interior blanks are
kept because a column's index is the only thing binding a value to a label; trailing
blanks are dropped because they bind nothing.
"""

from __future__ import annotations

import io
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from deepclare.intake import RoutedDocument
from deepclare.reading.errors import ReadingError

SHEET_HEADING = "=== Sheet: {name} ==="
"""How the whole-text rendering announces each sheet. One section per non-empty sheet, in
workbook order."""

BLANK_CELL = "(blank)"
"""How an empty cell is stated to a model. Absence is written out rather than rendered as
nothing: an omitted column and an empty one are different signals, and the labeller has to
be able to tell them apart."""

_GROUP_SEPARATORS = (" ", "\u00a0", "\u202f", "'", "\u2019")
"""Characters that appear inside a written number and mean nothing: ordinary and
non-breaking spaces, the narrow no-break space Excel emits for grouped numbers, and the
Swiss apostrophe."""


class Cell(BaseModel):
    """One cell, as the file stores it and as it renders.

    `number` is set only when the workbook itself holds a number there. A cell holding the
    *text* "1 250,50" has a number nobody has decided on yet — deciding is the typed
    reader's job, and it is allowed to refuse.
    """

    model_config = ConfigDict(frozen=True)

    text: str
    number: Decimal | None = None

    @property
    def is_blank(self) -> bool:
        return not self.text.strip()

    @property
    def is_short_text(self) -> bool:
        """A string of stripped length in (0, 40) — what a header cell looks like.

        Language-blind on purpose: this counts shape, never words.
        """
        return self.number is None and 0 < len(self.text.strip()) < 40


class SheetRow(BaseModel):
    """One non-empty row. `cells[i]` is spreadsheet column `i + 1`."""

    model_config = ConfigDict(frozen=True)

    number: int = Field(ge=1)
    """1-based row number in the sheet, kept so a note can name a cell a human can find."""

    cells: tuple[Cell, ...] = Field(min_length=1)

    def at(self, index: int) -> Cell | None:
        """The cell at a 0-based column index, or None when the row is not that wide."""
        if 0 <= index < len(self.cells):
            return self.cells[index]
        return None

    @property
    def first_cell(self) -> Cell:
        return self.cells[0]


class Sheet(BaseModel):
    """One sheet's non-empty rows, in order. A sheet with none is not buffered at all."""

    model_config = ConfigDict(frozen=True)

    name: str
    rows: tuple[SheetRow, ...] = Field(min_length=1)


class WorkbookBuffer(BaseModel):
    """Every non-empty sheet of one workbook, in workbook order."""

    model_config = ConfigDict(frozen=True)

    sheets: tuple[Sheet, ...] = Field(min_length=1)


def buffer_workbook(document: RoutedDocument) -> WorkbookBuffer:
    """Buffer every non-empty row of every sheet in one forward-only pass.

    Raises ReadingError when the file will not open as a workbook, or when it opens and
    holds no non-blank cell anywhere — a workbook with nothing in it is not an invoice
    that happens to be empty.
    """
    try:
        book = openpyxl.load_workbook(
            io.BytesIO(document.content), read_only=True, data_only=True
        )
    except (InvalidFileException, OSError, KeyError, ValueError) as exc:
        raise ReadingError(
            f"{document.file_name}: the file passed the workbook format check but will "
            f"not open as one: {exc}"
        ) from exc

    try:
        sheets = tuple(_buffer_sheet(book[name]) for name in book.sheetnames)
    except (OSError, ValueError, TypeError) as exc:
        raise ReadingError(
            f"{document.file_name}: the workbook opened but could not be read through: "
            f"{exc}"
        ) from exc
    finally:
        book.close()

    populated = tuple(sheet for sheet in sheets if sheet is not None)
    if not populated:
        raise ReadingError(
            f"{document.file_name}: the workbook has no non-blank cell in any of its "
            f"{len(sheets)} sheet(s). There is nothing to read."
        )
    return WorkbookBuffer(sheets=populated)


def render_workbook_text(buffer: WorkbookBuffer) -> str:
    """The whole workbook as tab-separated text, one section per sheet.

    Trailing blanks are already gone from the buffer; interior blanks render as empty
    fields so that a value stays under its header. Tabs and line breaks inside a cell
    become single spaces, because they would otherwise invent a column or a row.
    """
    sections = []
    for sheet in buffer.sheets:
        lines = [SHEET_HEADING.format(name=sheet.name)]
        lines.extend(
            "\t".join(_flatten(cell.text) for cell in row.cells) for row in sheet.rows
        )
        sections.append("\n".join(lines))
    return "\n\n".join(sections)


def render_row(row: SheetRow | None, column_count: int) -> str:
    """One row as tab-separated text, padded to `column_count`, blanks stated.

    Used for the column labeller's payload, where a blank cell has to be visible: a
    trailing run of empty fields is invisible to a reader and the labeller is asked to
    account for every column index it was given. `None` is a table whose header row is
    the top of the sheet — every column is blank, and the labeller judges on values.
    """
    cells = row.cells if row is not None else ()
    return "\t".join(
        BLANK_CELL
        if index >= len(cells) or cells[index].is_blank
        else _flatten(cells[index].text)
        for index in range(column_count)
    )


def parse_number(text: str) -> Decimal | None:
    """Read a written number, or refuse.

    Refusing matters more than parsing. This is the path the specification records as a
    silent failure — an unparseable cell became an empty field with no error and no log —
    and every caller here is required to say so out loud. So an ambiguous form is refused
    rather than guessed: `1,234` is 1234 in one convention and 1.234 in another, and a
    wrong weight on a customs declaration is worse than a missing one.
    """
    stripped = text.strip()
    for separator in _GROUP_SEPARATORS:
        stripped = stripped.replace(separator, "")

    sign = ""
    if stripped[:1] in ("+", "-"):
        sign, stripped = ("-" if stripped[0] == "-" else ""), stripped[1:]
    if not any(character.isdigit() for character in stripped):
        return None

    dots, commas = stripped.count("."), stripped.count(",")
    if dots and commas:
        # The rightmost of the two is the decimal mark; the other groups digits.
        decimal_mark = "." if stripped.rfind(".") > stripped.rfind(",") else ","
        stripped = stripped.replace("." if decimal_mark == "," else ",", "")
        stripped = stripped.replace(decimal_mark, ".")
    elif dots or commas:
        mark = "." if dots else ","
        head, _, tail = stripped.rpartition(mark)
        if stripped.count(mark) > 1:
            stripped = stripped.replace(mark, "")
        elif head and len(tail) == 3:
            # `1,234` / `1.234`: a thousands group and a three-decimal fraction are
            # written identically. Nothing in the cell says which, so nothing decides.
            return None
        else:
            stripped = f"{head}.{tail}" if head else f"0.{tail}"

    if not stripped or not stripped.replace(".", "").isdigit():
        return None
    try:
        return Decimal(f"{sign}{stripped}")
    except InvalidOperation:
        return None


def _buffer_sheet(sheet: Any) -> Sheet | None:
    """One sheet's non-empty rows. None when the sheet holds nothing."""
    # The declared dimensions are discarded before a single row is read. Whole-column
    # formatting inflates them to the sheet maximum, and the reader pads every row out to
    # that width, so this call is the difference between the real table and a rectangle of
    # phantom cells.
    sheet.reset_dimensions()

    rows: list[SheetRow] = []
    for cells in sheet.iter_rows():
        buffered = _buffer_row(cells)
        if buffered is not None:
            rows.append(buffered)
    return Sheet(name=sheet.title, rows=tuple(rows)) if rows else None


def _buffer_row(cells: tuple[Any, ...]) -> SheetRow | None:
    """One row, trailing blanks dropped. None when nothing in it is printed."""
    rendered = [_cell(getattr(cell, "value", None)) for cell in cells]
    while rendered and rendered[-1].is_blank:
        rendered.pop()
    if not rendered:
        return None

    # A row openpyxl yields is a mix of real cells and gap placeholders, and only a real
    # cell knows its row number. Every buffered row has at least one.
    number = next((cell.row for cell in cells if hasattr(cell, "row")), None)
    if number is None:
        return None
    return SheetRow(number=number, cells=tuple(rendered))


def _cell(value: object) -> Cell:
    """What one stored value is, and what it reads as."""
    match value:
        case None:
            return Cell(text="")
        case bool():
            return Cell(text="TRUE" if value else "FALSE")
        case int() | float() | Decimal():
            number = Decimal(str(value))
            return Cell(text=_plain(number), number=number)
        case datetime():
            # A date cell stores a serial number and a display format, so there is no
            # printed form to preserve. ISO 8601 is stated rather than guessed at.
            return Cell(
                text=value.date().isoformat()
                if value.time() == time()
                else value.isoformat(sep=" ")
            )
        case date() | time():
            return Cell(text=value.isoformat())
        case _:
            return Cell(text=str(value))


def _plain(number: Decimal) -> str:
    """A number written out in full: no exponent, no trailing zeros, no separators."""
    normalized = number.normalize()
    return format(normalized, "f")


def _flatten(text: str) -> str:
    """Collapse the two characters that would invent structure in a tab-separated dump."""
    return text.replace("\t", " ").replace("\r", " ").replace("\n", " ")
