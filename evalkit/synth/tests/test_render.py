"""Smoke tests for the document renderers (skipped if Pillow is absent)."""

from __future__ import annotations

import io
import random
import unittest

import openpyxl

try:
    from PIL import Image  # noqa: F401

    HAVE_PIL = True
except ImportError:
    HAVE_PIL = False

from synth.recombine import make_case
from synth.seed import Seed
from synth.tests.test_synth import _SEED_XML, load_seed_from_string


def _case():
    return make_case(load_seed_from_string(_SEED_XML), 3)


@unittest.skipUnless(HAVE_PIL, "Pillow not installed")
class RenderTests(unittest.TestCase):
    def test_invoice_pdf_is_a_pdf(self) -> None:
        from synth.render_docs import render_invoice_pdf

        data = render_invoice_pdf(_case(), random.Random(1))
        self.assertTrue(data.startswith(b"%PDF"))
        self.assertGreater(len(data), 2000)

    def test_cmr_pdf_is_a_pdf(self) -> None:
        from synth.render_docs import render_cmr_pdf

        self.assertTrue(render_cmr_pdf(_case(), random.Random(1)).startswith(b"%PDF"))

    def test_invoice_image_is_a4(self) -> None:
        from synth.paper import A4
        from synth.render_docs import invoice_image

        self.assertEqual(invoice_image(_case(), random.Random(1)).size, A4)

    def test_xlsx_opens_and_has_rows(self) -> None:
        from synth.render_docs import render_invoice_xlsx

        wb = openpyxl.load_workbook(io.BytesIO(render_invoice_xlsx(_case())))
        self.assertGreaterEqual(wb.active.max_row, 6)
        self.assertEqual(wb.active.max_column, 9)  # + Net kg column


class SeedTypeTests(unittest.TestCase):
    def test_seed_is_the_expected_type(self) -> None:
        self.assertIsInstance(load_seed_from_string(_SEED_XML), Seed)


if __name__ == "__main__":
    unittest.main()
