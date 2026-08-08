"""The spreadsheet path, A7 to A11, with the provider stubbed.

Nothing here touches the network. The two model calls answer from a stub transport, so
what is asserted is the payload that would have gone out and the record built from an
answer — including every degradation the specification prescribes for this path.
"""

from __future__ import annotations

import io
import json
from decimal import Decimal
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook

from deepclare.config import Settings
from deepclare.domain import DocumentRole
from deepclare.intake import RoutedDocument
from deepclare.intake.formats import FileFormat
from deepclare.models import GenerativeModel
from deepclare.reading import ReadingError
from deepclare.reading.schemas import LabelColumns, ReadWorkbookInvoice
from deepclare.reading.spreadsheet import read_workbook_invoice
from deepclare.reading.table import locate_goods_table
from deepclare.reading.workbook import (
    BLANK_CELL,
    buffer_workbook,
    parse_number,
    render_workbook_text,
)

PROMPTS = Path(__file__).resolve().parents[1] / "prompts"

HEADER = ["No", "Description", "Qty", "Unit", "Gross kg", "Net kg"]
GOODS = [
    [1, "POLYETHYLENE BAG 50X80", 12000, "PCS", 318.5, 300.0],
    [2, "STRETCH FILM 500MM", 240, "ROLL", 1104.0, 1056.0],
]


def settings(prompts_dir: Path) -> Settings:
    return Settings(
        google_api_key="not-a-real-key",
        genai_api_base="https://example.invalid/v1beta",
        genai_model_cheap="cheap",
        genai_model_standard="standard",
        genai_model_strong="strong",
        genai_max_output_tokens=1024,
        genai_timeout_seconds=30.0,
        prompts_dir=prompts_dir,
        reference_tables_dir=prompts_dir,
        qdrant_path=prompts_dir,
        qdrant_collection="codes",
        reference_dir=prompts_dir,
        reference_snapshot_dir=prompts_dir,
        nomenclature_api_base="https://example.invalid",
        nomenclature_max_node_id=1,
        nomenclature_crawl_workers=1,
    )


def workbook(
    preamble: list[tuple[str, object]] | None = None,
    header: list[object] | None = None,
    goods: list[list[object]] | None = None,
    totals: list[object] | None = None,
    first_row: int = 5,
    extra_sheet: dict[str, str] | None = None,
) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "Invoice"
    for reference, value in preamble or [("A1", "COMMERCIAL INVOICE"), ("A2", "SELLER")]:
        sheet[reference] = value
    if header is not None:
        for column, title in enumerate(header, start=1):
            sheet.cell(row=first_row, column=column, value=title)
    for offset, row in enumerate(goods if goods is not None else GOODS):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=first_row + 1 + offset, column=column, value=value)
    if totals:
        for column, value in enumerate(totals, start=1):
            sheet.cell(row=first_row + 1 + len(goods or GOODS), column=column, value=value)
    if extra_sheet:
        other = book.create_sheet(next(iter(extra_sheet)))
        other["A1"] = next(iter(extra_sheet.values()))
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def document(content: bytes) -> RoutedDocument:
    return RoutedDocument(
        document_id="doc1",
        file_name="invoice.xlsx",
        role=DocumentRole.INVOICE,
        role_was_declared=True,
        file_format=FileFormat.WORKBOOK,
        content=content,
    )


def whole_text_answer(**fields: object) -> ReadWorkbookInvoice:
    defaults: dict[str, object] = {
        "invoice_number": {"text": "MPS-2026-0417", "confidence": 0.95},
        "currency": {"text": "EUR", "confidence": 0.9},
        "goods_lines": [
            {"description": "GUESSED LINE", "quantity": 1, "confidence": 0.4}
        ],
    }
    return ReadWorkbookInvoice(**(defaults | fields))


def labels(*pairs: tuple[int, str]) -> LabelColumns:
    return LabelColumns(columns=[{"column": c, "label": l} for c, l in pairs])


DEFAULT_LABELS = labels(
    (0, "printed_line_number"),
    (1, "description"),
    (2, "quantity"),
    (3, "unit"),
    (4, "gross_weight"),
    (5, "net_weight"),
)


def model_over(
    sent: list[httpx.Request],
    whole_text: ReadWorkbookInvoice | None = None,
    labelling: LabelColumns | None = None,
    labelling_status: int = 200,
) -> GenerativeModel:
    """A provider that answers the whole-text read, then every labelling call."""
    answers = [(whole_text or whole_text_answer(), 200)]

    def respond(request: httpx.Request) -> httpx.Response:
        sent.append(request)
        if len(sent) == 1:
            body, status = answers[0]
        else:
            body, status = (labelling or DEFAULT_LABELS), labelling_status
        if status != 200:
            return httpx.Response(status, text="the labeller is unavailable")
        return httpx.Response(
            200,
            json={
                "candidates": [
                    {
                        "content": {"parts": [{"text": body.model_dump_json()}]},
                        "finishReason": "STOP",
                    }
                ],
                "modelVersion": "cheap-001",
            },
        )

    return GenerativeModel(
        settings(PROMPTS), httpx.Client(transport=httpx.MockTransport(respond))
    )


def prompt_text(request: httpx.Request) -> str:
    parts = json.loads(request.content)["contents"][0]["parts"]
    return "".join(part["text"] for part in parts if "text" in part)


# --- A7: the streaming load -----------------------------------------------------------


def test_only_non_blank_rows_are_buffered_and_trailing_blanks_are_dropped() -> None:
    sheet = buffer_workbook(document(workbook(header=HEADER))).sheets[0]
    # Rows 3 and 4 are blank and are not buffered; the preamble, the header at 5 and the
    # two goods rows are.
    assert [row.number for row in sheet.rows] == [1, 2, 5, 6, 7]
    assert [cell.text for cell in sheet.rows[0].cells] == ["COMMERCIAL INVOICE"]


def test_an_interior_blank_keeps_every_later_value_in_its_own_column() -> None:
    goods = [[1, "BAG", None, "PCS", 318.5, 300.0]]
    sheet = buffer_workbook(document(workbook(header=HEADER, goods=goods))).sheets[0]
    data = sheet.rows[-1]
    assert data.at(2) is not None and data.at(2).is_blank
    assert data.at(3).text == "PCS"
    assert data.at(4).number == Decimal("318.5")


def test_a_workbook_with_nothing_in_it_is_a_reading_failure() -> None:
    book = Workbook()
    buffer = io.BytesIO()
    book.save(buffer)
    with pytest.raises(ReadingError, match="no non-blank cell"):
        buffer_workbook(document(buffer.getvalue()))


def test_the_text_rendering_names_each_sheet_and_separates_cells_with_tabs() -> None:
    text = render_workbook_text(
        buffer_workbook(
            document(workbook(header=HEADER, extra_sheet={"Notes": "packed by shipper"}))
        )
    )
    assert "=== Sheet: Invoice ===" in text
    assert "=== Sheet: Notes ===" in text
    assert "No\tDescription\tQty\tUnit\tGross kg\tNet kg" in text


def test_an_ambiguous_written_number_is_refused_rather_than_guessed() -> None:
    assert parse_number("1 250,50") == Decimal("1250.50")
    assert parse_number("1.234,56") == Decimal("1234.56")
    # 1234, or 1.234? Nothing in the cell says, so nothing here decides.
    assert parse_number("1,234") is None
    assert parse_number("about 500") is None


# --- A9: locating the table -----------------------------------------------------------


def test_the_numbered_run_finds_the_header_and_excludes_the_totals_row() -> None:
    content = workbook(header=HEADER, totals=["", "TOTAL", "", "", 1422.5, 1356.0])
    sheet = buffer_workbook(document(content)).sheets[0]
    location = locate_goods_table(sheet)
    assert location is not None
    assert location.located_by == "numbered_run"
    assert sheet.rows[location.header_index].at(0).text == "No"
    assert [row.number for row in sheet.rows[location.first_data_index : location.last_data_index + 1]] == [6, 7]


def test_without_a_numbered_run_the_most_header_like_row_wins() -> None:
    goods = [["A", "POLYETHYLENE BAG", 12000], ["B", "STRETCH FILM", 240]]
    content = workbook(header=["Ref", "Description", "Qty"], goods=goods)
    sheet = buffer_workbook(document(content)).sheets[0]
    location = locate_goods_table(sheet)
    assert location is not None
    assert location.located_by == "header_score"
    assert sheet.rows[location.header_index].at(0).text == "Ref"


def test_a_sheet_with_no_table_locates_nothing_and_that_is_not_an_error() -> None:
    content = workbook(header=None, goods=[], preamble=[("A1", "Notes"), ("A2", "none")])
    sheet = buffer_workbook(document(content)).sheets[0]
    assert locate_goods_table(sheet) is None


# --- A10: labelling the columns -------------------------------------------------------


def test_the_labeller_is_sent_the_header_and_three_sample_rows_and_no_more() -> None:
    sent: list[httpx.Request] = []
    goods = [[n, f"ITEM {n}", n * 10, "PCS", 1.0, 0.9] for n in range(1, 6)]
    read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)),
        model_over(sent),
        PROMPTS,
    )
    payload = prompt_text(sent[1])
    assert "No\tDescription\tQty\tUnit\tGross kg\tNet kg" in payload
    assert "ITEM 1" in payload and "ITEM 3" in payload
    assert "ITEM 4" not in payload


def test_a_blank_header_cell_is_stated_rather_than_rendered_as_nothing() -> None:
    sent: list[httpx.Request] = []
    read_workbook_invoice(
        document(workbook(header=["No", None, "Qty", "Unit", "Gross kg", "Net kg"])),
        model_over(sent),
        PROMPTS,
    )
    assert BLANK_CELL in prompt_text(sent[1])


def test_two_columns_labelled_the_same_field_leave_the_lowest_index_standing() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)),
        model_over([], labelling=labels(
            (0, "printed_line_number"),
            (1, "description"),
            (4, "net_weight"),
            (5, "net_weight"),
        )),
        PROMPTS,
    )
    duplicates = reading.sheets[0].labelling.duplicates
    assert len(duplicates) == 1
    assert duplicates[0].field == "net_weight"
    assert duplicates[0].kept_column == 4
    assert duplicates[0].dropped_columns == (5,)
    # Column 4 holds the gross weight in this sheet; the wrong binding is what got filed.
    assert reading.reading.invoice.goods_lines[0].net_weight.value == Decimal("318.5")


def test_a_column_the_table_does_not_have_is_dropped_not_read() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)),
        model_over([], labelling=labels((1, "description"), (11, "quantity"))),
        PROMPTS,
    )
    assert reading.sheets[0].labelling.out_of_range_columns == (11,)
    assert reading.reading.invoice.goods_lines[0].quantity is None


def test_a_failed_labelling_call_falls_back_to_the_whole_text_guess() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)),
        model_over([], labelling_status=503),
        PROMPTS,
    )
    assert reading.goods_source == "whole_text_guess"
    assert reading.labelling_failure is not None
    assert [line.description.value for line in reading.reading.invoice.goods_lines] == [
        "GUESSED LINE"
    ]


# --- A11: reading the typed cells -----------------------------------------------------


def test_the_typed_cells_replace_the_guess_and_keep_both_weights_apart() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)), model_over([]), PROMPTS
    )
    assert reading.goods_source == "typed_cells"
    lines = reading.reading.invoice.goods_lines
    assert [line.line_id for line in lines] == ["1", "2"]
    assert lines[0].description.value == "POLYETHYLENE BAG 50X80"
    assert lines[0].gross_weight.value == Decimal("318.5")
    assert lines[0].net_weight.value == Decimal("300")
    assert lines[0].quantity.value == Decimal("12000")
    assert lines[0].unit.value == "PCS"
    assert lines[0].printed_line_number.value == 1


def test_the_header_fields_come_from_the_whole_text_read_even_so() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)), model_over([]), PROMPTS
    )
    assert reading.reading.invoice.invoice_number.value == "MPS-2026-0417"
    assert reading.reading.invoice.currency.value == "EUR"


def test_a_row_with_no_description_is_skipped_rather_than_filed_blank() -> None:
    goods = [
        [1, "POLYETHYLENE BAG", 12000, "PCS", 318.5, 300.0],
        [2, None, 0, None, 0, 0],
        [3, "STRETCH FILM", 240, "ROLL", 1104.0, 1056.0],
    ]
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)), model_over([]), PROMPTS
    )
    assert [line.line_id for line in reading.reading.invoice.goods_lines] == ["1", "2"]
    assert reading.sheets[0].rows_without_description == (7,)


def test_a_sheet_whose_mapping_has_no_description_column_yields_nothing() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)),
        model_over([], labelling=labels((2, "quantity"), (4, "gross_weight"))),
        PROMPTS,
    )
    assert reading.goods_source == "whole_text_guess"


def test_a_cell_that_will_not_read_as_a_number_is_reported_and_not_swallowed() -> None:
    goods = [
        [1, "POLYETHYLENE BAG", "about 12000", "PCS", 318.5, 300.0],
        [2, "STRETCH FILM", "n/a", "ROLL", 1104.0, 1056.0],
    ]
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)), model_over([]), PROMPTS
    )
    assert reading.reading.invoice.goods_lines[0].quantity is None
    assert len(reading.unread_numbers) == 1
    unread = reading.unread_numbers[0]
    assert unread.column == 2
    assert unread.field_name == "quantity"
    assert unread.rows == (6, 7)
    assert unread.examples == ("about 12000", "n/a")


def test_a_freight_row_inside_the_table_is_not_also_filed_as_goods() -> None:
    """The one judgement the typed reader defers to the whole-text read.

    A structural pass cannot tell a freight row from a goods row — it is numbered,
    priced and shaped like one — so a row A8 named verbatim as a service charge is left
    out of the goods rather than filed twice.
    """
    goods = [
        [1, "POLYETHYLENE BAG", 12000, "PCS", 318.5, 300.0],
        [2, "FREIGHT MERSIN-YEREVAN", 1, "SERVICE", None, None],
    ]
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)),
        model_over(
            [],
            whole_text=whole_text_answer(
                service_charges=[
                    {
                        "description": "  freight mersin-yerevan ",
                        "amount": 620,
                        "confidence": 0.8,
                    }
                ]
            ),
        ),
        PROMPTS,
    )
    assert [line.description.value for line in reading.reading.invoice.goods_lines] == [
        "POLYETHYLENE BAG"
    ]
    assert reading.sheets[0].service_rows == (7,)
    assert reading.reading.service_charges[0].amount.value == Decimal("620")


def test_a_service_row_the_whole_text_read_paraphrased_stays_as_goods() -> None:
    goods = [
        [1, "POLYETHYLENE BAG", 12000, "PCS", 318.5, 300.0],
        [2, "FREIGHT MERSIN-YEREVAN", 1, "SERVICE", None, None],
    ]
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)),
        model_over(
            [],
            whole_text=whole_text_answer(
                service_charges=[
                    {"description": "transport costs", "confidence": 0.5}
                ]
            ),
        ),
        PROMPTS,
    )
    assert len(reading.reading.invoice.goods_lines) == 2
    assert reading.sheets[0].service_rows == ()


def test_an_empty_cell_is_not_reported_as_unreadable() -> None:
    goods = [[1, "POLYETHYLENE BAG", None, "PCS", 318.5, 300.0]]
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, goods=goods)), model_over([]), PROMPTS
    )
    assert reading.unread_numbers == ()


def test_a_sheet_with_no_table_contributes_nothing_and_raises_nothing() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER, extra_sheet={"Notes": "packed by the shipper"})),
        model_over([]),
        PROMPTS,
    )
    assert [sheet.sheet_name for sheet in reading.sheets] == ["Invoice", "Notes"]
    assert reading.sheets[1].table is None
    assert reading.sheets[1].goods_rows == 0
    assert reading.goods_source == "typed_cells"


def test_a_workbook_value_is_extracted_and_names_the_prompt_that_bound_it() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)), model_over([]), PROMPTS
    )
    line = reading.reading.invoice.goods_lines[0]
    assert line.description.provenance.source_document_id == "doc1"
    assert line.description.provenance.prompt_name == "label_columns"
    assert line.description.provenance.region is None
    assert reading.reading.invoice.currency.provenance.prompt_name == (
        "read_workbook_invoice"
    )


def test_no_goods_from_either_path_is_a_reading_failure() -> None:
    with pytest.raises(ReadingError, match="nothing to declare"):
        read_workbook_invoice(
            document(workbook(header=HEADER)),
            model_over([], whole_text=whole_text_answer(goods_lines=[]),
                       labelling=labels((2, "quantity"))),
            PROMPTS,
        )


def test_the_run_records_every_call_it_made() -> None:
    reading = read_workbook_invoice(
        document(workbook(header=HEADER)), model_over([]), PROMPTS
    )
    assert [call.prompt_name for call in reading.model_calls] == [
        "read_workbook_invoice",
        "label_columns",
    ]
