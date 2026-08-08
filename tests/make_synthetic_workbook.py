"""Synthetic spreadsheet invoice, entirely fictitious, built to exercise the whole A7–A11
path rather than to look tidy.

Everything the spreadsheet path has to survive is in it on purpose:

* a **free-form preamble** of seven rows above the table, including a blank row and a
  right-hand block of cells, so the table does not start at row 1 and no fixed offset
  finds it;
* the goods table starting at **row 9**, its own row numbering in column A;
* a **totals row** directly under the data, numbered by nothing, which the numbered-run
  rule must exclude;
* **Armenian column headers**, which is the case the whole language-blind design exists
  for — the workbook channel is the only route in this system carrying Armenian goods
  text;
* **net weight beside gross weight**, in the order gross-then-net, which is the reproduced
  bug: one model transcribing the table drops one of the pair;
* a column headed `Կոդ` holding the seller's own article numbers, which must not be read
  as a customs code, and a separate `ԱՏԳ ԱԱ ԿՈԴ` column that must;
* a **quantity column where one cell holds text** (`շուրջ 500`, "about 500"), which is the
  cell that would silently become empty;
* one row with a **blank description**, which must be skipped rather than filed;
* a **freight row** inside the table, which is not goods; and
* a **second sheet with no table at all**, which must contribute nothing and raise
  nothing.

The declared used range is then inflated by hand to `A1:BZ50000`, which is what
whole-column formatting does to a real workbook, so the loader is measured against the
staleness it exists to survive.

    .venv/bin/python tests/make_synthetic_workbook.py
"""

from __future__ import annotations

import io
import re
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook

OUT = Path("/tmp/invoice_synthetic.xlsx")
INFLATED_DIMENSION = b'<dimension ref="A1:BZ50000"/>'

PREAMBLE = [
    ("A1", "ՀԱՇԻՎ-ԱՊՐԱՆՔԱԳԻՐ / COMMERCIAL INVOICE"),
    ("A2", "ԱՌԱՔՈՂ:"),
    ("B2", "MERSIN PLASTIK SANAYI A.S."),
    ("B3", "Ataturk Bulvari 214, 33100 Mersin, Turkiye"),
    ("B4", "VKN 4820996317"),
    ("E2", "Invoice No:"),
    ("F2", "MPS-2026-0417"),
    ("E3", "Date:"),
    ("E4", "Terms:"),
    ("F4", "FCA Mersin"),
    ("E5", "Currency:"),
    ("F5", "EUR"),
    ("A6", "ՍՏԱՑՈՂ:"),
    ("B6", 'ԱՐԱՐԱՏ ՓԱԹԵԹԱՎՈՐՈՒՄ ՍՊԸ'),
    ("B7", "Երևան, Արշակունյաց 87, ՀՀ"),
    ("E6", "TIN:"),
    ("F6", "02845519"),
]

HEADER = [
    "Հ/Հ",
    "Ապրանքի անվանումը",
    "Կոդ",
    "ԱՏԳ ԱԱ ԿՈԴ",
    "Քանակ",
    "Չափ. միավոր",
    "Համախառն քաշ, կգ",
    "Զուտ քաշ, կգ",
    "Միավորի գին",
    "Ընդամենը",
    "Ծագման երկիր",
]

# (number, description, article, hs, qty, unit, gross, net, unit price, total, origin)
GOODS = [
    (1, "ՊՈԼԻԷԹԻԼԵՆԱՅԻՆ ՊԱՐԿ 50X80 ՍՄ", "ART-5080", "3923210000",
     12000, "հատ", 318.5, 300.0, 0.042, 504.00, "TR"),
    (2, "ՊՈԼԻՊՐՈՊԻԼԵՆԱՅԻՆ ՊԱՐԿ 55X95 ՍՄ", "ART-5595", "3923290000",
     "շուրջ 500", "հատ", 92.4, 88.0, 0.115, 57.50, "TR"),
    (3, "ՍՏՐԵՉ ԹԱՂԱՆԹ 500ՄՄ X 300Մ", "ART-STR50", "3919109000",
     240, "ռուլոն", 1104.0, 1056.0, 8.75, 2100.00, "TR"),
    (4, "", "ART-XXXX", "", 0, "", 0, 0, 0, 0, ""),
    (5, "ԿՈՆՏԵՅՆԵՐԱՅԻՆ ՆԵՐԴԻՐ 20 ՖՈՒՏ", "ART-LIN20", "3923900000",
     60, "հատ", 471.0, 450.0, 14.20, 852.00, "TR"),
    (6, "ՓՈԽԱԴՐՄԱՆ ԾԱԽՍ / FREIGHT MERSIN-YEREVAN", "", "",
     1, "ծառայություն", 0, 0, 620.00, 620.00, ""),
]

TOTALS = ["", "ԸՆԴԱՄԵՆԸ", "", "", "", "", 1985.9, 1894.0, "", 4133.50, ""]


def build() -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "Invoice"

    for reference, value in PREAMBLE:
        sheet[reference] = value
    sheet["F3"] = date(2026, 3, 12)

    for column, title in enumerate(HEADER, start=1):
        sheet.cell(row=9, column=column, value=title)
    for offset, row in enumerate(GOODS):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=10 + offset, column=column, value=value or None)
    for column, value in enumerate(TOTALS, start=1):
        sheet.cell(row=17, column=column, value=value or None)

    notes = book.create_sheet("Notes")
    notes["A1"] = "Packing carried out under supervision of the shipper."
    notes["A3"] = "All goods of Turkish origin unless stated otherwise on the line."

    buffer = io.BytesIO()
    book.save(buffer)
    return _inflate_declared_range(buffer.getvalue())


def _inflate_declared_range(content: bytes) -> bytes:
    """Rewrite the sheet's declared dimension the way whole-column formatting does.

    openpyxl writes an honest dimension. A workbook that has had a fill or a border
    applied down a whole column does not, and that stale range is what a materializing
    loader pays for.
    """
    source = zipfile.ZipFile(io.BytesIO(content))
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data, count = re.subn(
                    rb'<dimension ref="[^"]*"\s*/>', INFLATED_DIMENSION, data
                )
                if count != 1:
                    raise RuntimeError(
                        "the sheet XML has no single <dimension> element to inflate; "
                        "the point of this fixture is the stale range"
                    )
            target.writestr(item, data)
    return out.getvalue()


if __name__ == "__main__":
    OUT.write_bytes(build())
    print(f"wrote {OUT} ({OUT.stat().st_size} bytes)")
